import sqlite3
import is_login
from datetime import datetime
import csv
import os
import schedule
import time
from openpyxl import Workbook, load_workbook



def record_of_login_in_csv_file(student_username):
    try:

        current_time = datetime.now().strftime("%y-%m-%d %H:%M:%S")
        with open('user_login.csv', 'a') as csv_file:
            csv_write = csv.writer(csv_file)
            csv_write.writerow([student_username, current_time])
            print("\ndata sucessfully saved in csv file")
    except Exception as e:
        print("\nError occurred while saving data to CSV:", e)




def transfer_csv_data_from_database_to_excel():

    try:
        excel_path = 'login_history.xlsx'
        if not os.path.exists(excel_path):

            workbook = Workbook()
            worksheet = workbook.active
            worksheet['A1'] = "Year"
            worksheet['B1'] = "Month"
            worksheet['C1'] = "Day"
            worksheet['D1'] = "Hour"
            worksheet['E1'] = "Login_Count"
            workbook.save(excel_path)
        else:

            workbook = load_workbook(excel_path)
            worksheet = workbook.active

        login_counts_per_hour = {}

        with open('user_login.csv', 'r') as csv_file:
            csv_reader = csv.reader(csv_file)

            for row in csv_reader:

                username, timestamp_str = row
                timestamp = datetime.strptime(timestamp_str, "%y-%m-%d %H:%M:%S")
                year = timestamp.year
                month = timestamp.month
                day = timestamp.day
                hour = timestamp.hour

                login_counts_per_hour[(year, month, day, hour)] = login_counts_per_hour.get((year, month, day, hour), 0) + 1

        for (year, month, day, hour), login_count in login_counts_per_hour.items():

            next_row = worksheet.max_row + 1

            worksheet.cell(row=next_row, column=1, value=year)
            worksheet.cell(row=next_row, column=2, value=month)
            worksheet.cell(row=next_row, column=3, value=day)
            worksheet.cell(row=next_row, column=4, value=hour)
            worksheet.cell(row=next_row, column=5, value=login_count)


        workbook.save(excel_path)
        print("\nData successfully transferred to Excel file.")
    except Exception as e:
        print("\nError occurred while transferring CSV data to Excel:", e)


def transfer_csv_data_to_inloggning_database():
    try:
    
        with open('user_login.csv', 'r') as file:
            csv_reader = csv.reader(file)
            
            conn = sqlite3.connect('inloggning_database.db')
            cursor = conn.cursor()

        
            cursor.execute('''CREATE TABLE IF NOT EXISTS inloggning (
                                user_name TEXT,
                            current_time DATETIME
                                )''')
            for row in csv_reader:
                cursor.execute('INSERT INTO inloggning  VALUES (?, ?)', row)
            
                
            conn.commit()
            conn.close()
        
        print("Data transferred to SQLite successfully.")

    
        with open('user_login.csv', 'w', newline='') as file:
            pass

        print("CSV file cleaned successfully.")
    except Exception as e:
        print("\nError occurred while transferring CSV data to SQLite database:", e)



def schedule_task():
    try:

        schedule.every().hour.do(transfer_csv_data_from_database_to_excel)
        schedule.every().hour.do(transfer_csv_data_to_inloggning_database)


        while True:
            schedule.run_pending()
            time.sleep(1)
    except Exception as e:
        print("\nError occurred in schedule task:", e)




def student_login():
    try:
        conn = sqlite3.connect('student_database.db')
        cur = conn.cursor()

        student_username = input(" enter username: ")

        cur.execute('''select * from student where
                    student_username = ? ''', (student_username,))
        user = cur.fetchone()

        if user:
            password = input("enter your password: ")
            cur.execute(''' select password from student where
                        student_username = ?''',(student_username,))
            save_password = cur.fetchone()

            if save_password and password == save_password[0]:
                print("login sucessfully")
                record_of_login_in_csv_file(student_username)
                is_login.is_logged_in(student_username)

            else:
                print("fail password")
        else:
            print("user not exists")
        
        
        conn.commit()
        conn.close()
    
    except Exception as e:
        print("\nError occurred during student login:", e)
        